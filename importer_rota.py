"""
OncoScheduler v2 — ROTA importer.

Parses the secretary's weekly files:
  1. ROTA xlsx  (Fellows-Rotation-Residents sheet): roster, leaves, LTC pins,
     excused sessions, resident rotation windows.
  2. Clinics xlsx (patient bookings): Day | Session | Consultant | Specialty | #Patients.

Design: the importer NEVER writes to the database directly. It returns a list
of proposed changes; the GUI shows them and the chief confirms which to apply.
Unparseable remarks are surfaced as 'review' items, never silently dropped.
"""

import re
from datetime import date, datetime, timedelta

import openpyxl

# Rotations that mean "in our clinics this period"
CLINIC_ROTATIONS = {
    "gi/gu": "GI/GU",
    "breast": "Breast",
    "gyn/lung": "Gyne/Lung",
    "gyne/lung": "Gyne/Lung",
    "sarcoma": "Sarcoma",
    "sarcoma/h&n": "Sarcoma",
}

DAY_TOKENS = {
    "sun": "sun", "sunday": "sun",
    "mon": "mon", "monday": "mon",
    "tue": "tue", "tues": "tue", "tuesday": "tue",
    "wed": "wed", "wednesday": "wed",
    "thu": "thu", "thur": "thu", "thurs": "thu", "thursday": "thu",
}

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}


def _norm(s):
    return str(s or "").strip()


def _clean_name(raw):
    """'Dr. Gazi Sindi' -> 'Gazi Sindi'; strips the Dr. prefix and whitespace."""
    n = _norm(raw)
    n = re.sub(r"^\s*dr\.?\s+", "", n, flags=re.I)
    return re.sub(r"\s+", " ", n).strip()


def first_name(full):
    """Roster short name = first token (matches how Fawaz's files name people)."""
    return _clean_name(full).split(" ")[0] if _clean_name(full) else ""


def parse_leave_range(text, year):
    """
    '01-18 June' / '01-18 June 2026' / '5-9 July' -> (date, date) or None.
    A bare month name -> the whole month.
    """
    t = _norm(text)
    if not t:
        return None
    m = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})\s+([A-Za-z]+)(?:\s+(\d{4}))?", t)
    if m:
        d1, d2, mon, yr = int(m.group(1)), int(m.group(2)), m.group(3).lower(), m.group(4)
        if mon in MONTHS:
            y = int(yr) if yr else year
            try:
                return (date(y, MONTHS[mon], d1), date(y, MONTHS[mon], d2))
            except ValueError:
                return None
    m = re.search(r"^([A-Za-z]+)(?:\s+(\d{4}))?$", t)
    if m and m.group(1).lower() in MONTHS:
        y = int(m.group(2)) if m.group(2) else year
        mo = MONTHS[m.group(1).lower()]
        last = (date(y, mo % 12 + 1, 1) - timedelta(days=1)) if mo < 12 else date(y, 12, 31)
        return (date(y, mo, 1), last)
    return None


def parse_ltc(remark):
    """'LTC: RAUF MON-AM' -> ('rauf', 'mon', 'AM') or None."""
    m = re.search(r"LTC\s*:\s*([A-Za-z]+)\s+([A-Za-z]+)\s*[- ]\s*(AM|PM)", _norm(remark), re.I)
    if not m:
        return None
    cons, day_tok, sess = m.group(1), m.group(2).lower(), m.group(3).upper()
    day = DAY_TOKENS.get(day_tok)
    if not day:
        return None
    return (cons.capitalize(), day, sess)


ALL_SESSIONS = [(d, s) for d in ["sun", "mon", "tue", "wed", "thu"] for s in ("AM", "PM")]


def parse_whitelist_schedule(remark):
    """
    Detect the NP-style WHITELIST pattern:
        'Breast (Sun AM, Mon PM, Tue AM, Wed PM) Thursday-Excused'
    meaning: works ONLY the listed sessions. Returns
        (specialty, exclusions) where exclusions = complement of the whitelist,
    or None if the remark is not a whitelist.
    """
    t = _norm(remark)
    m = re.search(r"^([A-Za-z/&]+)\s*\(([^)]+)\)", t)
    if not m:
        return None
    specialty = m.group(1).strip()
    inner = m.group(2)
    whitelist = set()
    for mm in re.finditer(r"\b(Sun|Mon|Tue|Tues|Wed|Thu|Thur|Thurs|Sunday|Monday|Tuesday|Wednesday|Thursday)\s*[- ]?\s*(AM|PM)\b", inner, re.I):
        d = DAY_TOKENS.get(mm.group(1).lower())
        if d:
            whitelist.add((d, mm.group(2).upper()))
    if not whitelist:
        return None
    exclusions = [(d, s) for (d, s) in ALL_SESSIONS if (d, s) not in whitelist]
    return specialty, exclusions


def parse_excused(remark):
    """
    Best-effort extraction of recurring excused sessions from free-text remarks.
    Handles the patterns observed in real ROTA files:
        'Excused: Wednesday PM'        -> [(wed, PM)]
        'Every Tuesday PM excused'     -> [(tue, PM)]
        'Excused every Wednesday PM'   -> [(wed, PM)]
        'Thursday-Excused'             -> [(thu, ALL)]
        'Excused AM (Thursday)'        -> [(thu, AM)]
        'Excused PM (Thursday) Wednesday PM' -> [(thu, PM), (wed, PM)]
    Returns (parsed_list, leftover_flag). leftover_flag=True means the remark
    contained 'excus' but we couldn't fully parse it -> needs manual review.
    """
    t = _norm(remark)
    if not t or "excus" not in t.lower():
        return [], False
    found = []

    # Pattern: 'Excused AM (Thursday)' / 'Excused PM (Thursday)'
    for m in re.finditer(r"excused\s+(AM|PM)\s*\(\s*([A-Za-z]+)\s*\)", t, re.I):
        d = DAY_TOKENS.get(m.group(2).lower())
        if d:
            found.append((d, m.group(1).upper()))

    # Pattern: 'DAY PM' / 'DAY AM' anywhere ('Every Tuesday PM excused',
    # 'Excused: Wednesday PM', trailing 'Wednesday PM')
    for m in re.finditer(r"\b([A-Za-z]+day|Sun|Mon|Tue|Tues|Wed|Thu|Thur|Thurs)\s+(AM|PM)\b", t, re.I):
        d = DAY_TOKENS.get(m.group(1).lower())
        if d and (d, m.group(2).upper()) not in found:
            found.append((d, m.group(2).upper()))

    # Pattern: 'Thursday-Excused' / 'Thursday Excused' (whole day)
    for m in re.finditer(r"\b([A-Za-z]+day)\s*[- ]\s*excused", t, re.I):
        d = DAY_TOKENS.get(m.group(1).lower())
        if d and not any(x[0] == d for x in found):
            found.append((d, "ALL"))

    leftover = not found  # contained 'excus' but nothing parsed
    return found, leftover


def parse_rota(path, default_year=None):
    """
    Parse the ROTA workbook. Returns a list of proposal dicts:
      {action: 'add_person'|'leave'|'pin'|'exclusion'|'review',
       name, short_name, kind, rotation, start, end, day, session,
       consultant, note, detail}
    """
    if default_year is None:
        default_year = date.today().year
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    proposals = []
    section = None  # 'fellows' | 'residents'
    seen_short = {}  # short name -> full name (collision detection)
    for row in rows:
        cells = [_norm(c) for c in row]
        joined = " ".join(cells).strip().upper()
        if not joined:
            continue
        if joined.startswith("FELLOWS"):
            section = "fellows"
            continue
        if joined.startswith("RESIDENTS"):
            section = "residents"
            continue
        if cells[0].lower() in ("name", ""):
            continue
        if section is None:
            continue

        full_name = _clean_name(cells[0])
        if not full_name:
            continue
        short = first_name(full_name)
        # Name collision: two people with the same first name (e.g. two Ahmeds).
        # Disambiguate the LATER one with first+last so they stay distinct.
        if short in seen_short and seen_short[short] != full_name:
            parts = full_name.split(" ")
            short = parts[0] + (" " + parts[-1] if len(parts) > 1 else "")
        seen_short.setdefault(short, full_name)

        # Column positions per observed layout:
        # 0=Name 1=Level 2=ID 3=MCD 4=Email 5=Start 6=End 7=Memo 8=Leave 9=Rotation 10=Sponsor 11=Remarks 12=Status
        start_raw = row[5] if len(row) > 5 else None
        end_raw = row[6] if len(row) > 6 else None
        leave_raw = _norm(row[8]) if len(row) > 8 else ""
        rotation_raw = _norm(row[9]) if len(row) > 9 else ""
        remarks = _norm(row[11]) if len(row) > 11 else ""

        start_d = start_raw.date() if isinstance(start_raw, datetime) else None
        end_d = end_raw.date() if isinstance(end_raw, datetime) else None

        if section == "fellows":
            rot_key = rotation_raw.lower().strip()
            in_clinics = rot_key in CLINIC_ROTATIONS
            rotation = CLINIC_ROTATIONS.get(rot_key, "")
            kind = "fellow"
            off_service = not in_clinics

            # NP-style whitelist schedule in remarks overrides:
            # 'Breast (Sun AM, Mon PM, ...)' = works ONLY those sessions.
            wl = parse_whitelist_schedule(remarks)
            if wl:
                specialty, wl_exclusions = wl
                kind = "np"
                rotation = CLINIC_ROTATIONS.get(specialty.lower(), specialty)
                off_service = False
                proposals.append({
                    "action": "add_person", "name": short, "full_name": full_name,
                    "kind": kind, "rotation": rotation,
                    "start": start_d, "end": end_d,
                    "detail": f"NP-style schedule, works only listed sessions ({rotation})",
                    "off_service": False,
                })
                for d, s in wl_exclusions:
                    proposals.append({
                        "action": "exclusion", "name": short, "full_name": full_name,
                        "day": d, "session": s,
                        "detail": f"not on whitelist → excluded {d.upper()} {s}"})
            else:
                proposals.append({
                    "action": "add_person", "name": short, "full_name": full_name,
                    "kind": kind, "rotation": rotation,
                    "start": start_d, "end": end_d,
                    "detail": (f"rotation '{rotation_raw}'"
                               + ("" if in_clinics else " — OFF-SERVICE this period (not in our clinics)")),
                    "off_service": off_service,
                })
            if leave_raw:
                rng = parse_leave_range(leave_raw, default_year)
                if rng:
                    proposals.append({
                        "action": "leave", "name": short, "full_name": full_name,
                        "start": rng[0], "end": rng[1],
                        "detail": f"leave '{leave_raw}'"})
                else:
                    proposals.append({
                        "action": "review", "name": short, "full_name": full_name,
                        "detail": f"Unparsed leave text: '{leave_raw}'"})
        else:  # residents
            proposals.append({
                "action": "add_person", "name": short, "full_name": full_name,
                "kind": "resident", "rotation": "",
                "start": start_d, "end": end_d,
                "detail": f"rotation window {start_d} → {end_d}",
                "off_service": False,
            })

        # LTC pin (both sections)
        ltc = parse_ltc(remarks)
        if ltc:
            proposals.append({
                "action": "pin", "name": short, "full_name": full_name,
                "consultant": ltc[0], "day": ltc[1], "session": ltc[2],
                "detail": f"LTC pin {ltc[0]} {ltc[1].upper()}-{ltc[2]}"})

        # Excused sessions (skip if the remark was a whitelist — already handled)
        if section == "fellows" and parse_whitelist_schedule(remarks):
            excl, leftover = [], False
        else:
            excl, leftover = parse_excused(remarks)
        for d, s in excl:
            proposals.append({
                "action": "exclusion", "name": short, "full_name": full_name,
                "day": d, "session": s,
                "detail": f"excused {d.upper()} {s}"})
        if leftover or (remarks and not ltc and not excl
                        and "excus" not in remarks.lower()
                        and any(k in remarks.lower() for k in ("clinic", "nf", "week"))):
            proposals.append({
                "action": "review", "name": short, "full_name": full_name,
                "detail": f"Remark needs manual review: '{remarks}'"})

    return proposals


def parse_patient_numbers(path):
    """
    Parse the secretary's patient-bookings file:
      Day | Session | Consultant | Specialty | # Of Patients
    Returns list of clinic dicts ready for week_clinics.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(values_only=True):
        day = _norm(row[0]).lower()
        if day in ("", "day") or len(row) < 5:
            continue
        d = DAY_TOKENS.get(day, DAY_TOKENS.get(day[:3]))
        if not d:
            continue
        sess = _norm(row[1]).upper()
        if sess not in ("AM", "PM"):
            continue
        cons = _norm(row[2]).capitalize()
        spec = _norm(row[3])
        try:
            pts = int(row[4])
        except (TypeError, ValueError):
            pts = 0
        out.append({"day": d, "session": sess, "consultant": cons,
                    "specialty": spec, "patients": pts, "clinic_type": ""})
    return out


def diff_against_db(proposals, db):
    """
    Annotate proposals against current DB state so the GUI can show
    NEW / EXISTS / CHANGED. Returns the same list with 'status' added.
    """
    for p in proposals:
        if p["action"] == "add_person":
            existing = db.get_person_by_name(p["name"])
            if existing is None:
                p["status"] = "NEW"
            elif (existing["rotation"] or "") != (p["rotation"] or "") and p["rotation"]:
                p["status"] = f"CHANGED (rotation {existing['rotation']} → {p['rotation']})"
            else:
                p["status"] = "EXISTS"
        elif p["action"] == "leave":
            person = db.get_person_by_name(p["name"])
            if person is None:
                p["status"] = "PERSON NOT IN DB"
            else:
                dup = any(
                    lv["start_date"] == p["start"].strftime("%Y-%m-%d")
                    and lv["end_date"] == p["end"].strftime("%Y-%m-%d")
                    for lv in db.list_leaves(person["id"]))
                p["status"] = "EXISTS" if dup else "NEW"
        elif p["action"] == "pin":
            person = db.get_person_by_name(p["name"])
            if person is None:
                p["status"] = "PERSON NOT IN DB"
            else:
                dup = any(
                    rp["day"] == p["day"] and rp["session"] == p["session"]
                    and rp["consultant"].lower() == p["consultant"].lower()
                    for rp in db.list_recurring_pins(person["id"]))
                p["status"] = "EXISTS" if dup else "NEW"
        elif p["action"] == "exclusion":
            person = db.get_person_by_name(p["name"])
            if person is None:
                p["status"] = "PERSON NOT IN DB"
            else:
                dup = any(
                    re_["day"] == p["day"] and re_["session"] == p["session"]
                    for re_ in db.list_recurring_exclusions(person["id"]))
                p["status"] = "EXISTS" if dup else "NEW"
        else:
            p["status"] = "REVIEW"
    return proposals


def apply_proposal(p, db):
    """Apply one confirmed proposal to the database. Returns description string."""
    if p["action"] == "add_person":
        existing = db.get_person_by_name(p["name"])
        if existing:
            if p.get("rotation") and existing["rotation"] != p["rotation"]:
                db.update_person(existing["id"], rotation=p["rotation"])
                return f"Updated rotation for {p['name']} → {p['rotation']}"
            return f"{p['name']} already in roster"
        kind = p["kind"]
        active = 0 if p.get("off_service") else 1
        pid = db.add_person(p["name"], kind, p.get("rotation", ""),
                            min_clinics=0, max_clinics=5,
                            start_date=p.get("start"), end_date=p.get("end"),
                            notes=p.get("full_name", ""))
        if not active:
            db.update_person(pid, active=0)
        return f"Added {kind} {p['name']}" + (" (inactive, off-service)" if not active else "")
    if p["action"] == "leave":
        person = db.get_person_by_name(p["name"])
        if not person:
            return f"SKIPPED leave — {p['name']} not in roster"
        db.add_leave(person["id"], p["start"], p["end"], note="from ROTA")
        return f"Leave for {p['name']}: {p['start']} → {p['end']}"
    if p["action"] == "pin":
        person = db.get_person_by_name(p["name"])
        if not person:
            return f"SKIPPED pin — {p['name']} not in roster"
        db.add_recurring_pin(person["id"], p["day"], p["session"], p["consultant"])
        return f"LTC pin for {p['name']}: {p['consultant']} {p['day'].upper()} {p['session']}"
    if p["action"] == "exclusion":
        person = db.get_person_by_name(p["name"])
        if not person:
            return f"SKIPPED exclusion — {p['name']} not in roster"
        if p["session"] == "ALL":
            db.add_recurring_exclusion(person["id"], p["day"], "AM")
            db.add_recurring_exclusion(person["id"], p["day"], "PM")
        else:
            db.add_recurring_exclusion(person["id"], p["day"], p["session"])
        return f"Excused {p['name']}: {p['day'].upper()} {p['session']}"
    return f"Review item (no action): {p.get('detail','')}"
