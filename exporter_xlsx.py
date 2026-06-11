"""Excel exporter for the weekly schedule (live formulas)."""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CONSULTANT_CODES = {
    "alsayed":"AS","suleman":"KS","akhtar":"SA","alhussaini":"HH",
    "bazarbashi":"SB","anwar":"MA","alghabban":"AGA","alqahtani":"AQ",
    "almugbel":"FA","tweigieri":"TT","atallah":"JA","aljubran":"AA",
    "aljabrani":"AA","alzahrani":"AZ","sabah":"SAK","meshari":"MAZ",
    "rauf":"MR","aisha":"AIS","alyahya":"MAY","chemoassessment":"CA",
}


def export_xlsx(result, fellows_df, residents_df, output_path,
                period_start, period_end, date_map):
    HDR = PatternFill("solid", fgColor="1F4E78")
    SUMHDR = PatternFill("solid", fgColor="2E75B6")
    DAY = PatternFill("solid", fgColor="D9E1F2")
    CHEMO = PatternFill("solid", fgColor="FFF2CC")
    ADDON = PatternFill("solid", fgColor="FCE4B5")  # light orange — distinct from chemo & warning
    GREY = PatternFill("solid", fgColor="F2F2F2")
    THIN = Side(border_style="thin", color="999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    WHITE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    BOLD = Font(name="Calibri", size=10, bold=True)
    NORM = Font(name="Calibri", size=10)
    TITLE = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    HEAD = Font(name="Calibri", size=11, bold=True, color="1F4E78")

    def code_for(consultant):
        return CONSULTANT_CODES.get(consultant.strip().lower(), "")

    def helpers_str(s):
        if not s or str(s).strip() in ("—", "-", ""):
            return ""
        parts = [p.strip() for p in str(s).split(",") if p.strip()]
        return "/".join(p.upper() for p in parts)

    wb = Workbook()

    # ============ SHEET 1: SCHEDULE ============
    ws = wb.active
    ws.title = "Schedule"
    ws["A1"] = "MEDICAL ONCOLOGY"
    ws["A1"].font = TITLE
    ws["A1"].alignment = CENTER
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 22
    ws["A2"] = "OUTPATIENT AND INPATIENT SCHEDULE"
    ws["A2"].font = HEAD
    ws["A2"].alignment = CENTER
    ws.merge_cells("A2:H2")
    ws["A3"] = "Period:"
    ws["A3"].font = BOLD
    ws["B3"] = f"{period_start}  -  {period_end}"
    ws["B3"].font = BOLD
    ws["G3"] = f"Revised: {datetime.now().strftime('%d %b %Y')}"
    ws["G3"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells("G3:J3")

    # Warning banner — counts gaps so they can't be missed
    n_warnings = sum(1 for row in result["schedule"] if row.get("Warnings"))
    if n_warnings > 0:
        warn_msg = (
            f"⚠ {n_warnings} clinic(s) have warnings — see red rows below "
            f"(staffing tight; manual review needed)"
        )
        ws["A4"] = warn_msg
        ws["A4"].font = Font(name="Calibri", size=11, bold=True, color="C00000")
        ws["A4"].alignment = CENTER
        ws.merge_cells("A4:J4")
        ws.row_dimensions[4].height = 22

    HEADERS = ["Day","Date","Session","Code","Consultant","Specialty","Patients",
               "Required","Helpers","Warnings"]
    HDR_ROW = 5
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.font = WHITE; c.fill = HDR; c.alignment = CENTER; c.border = BORDER

    # Red fill for warning rows
    WARN_FILL = PatternFill("solid", fgColor="FCE4D6")  # light red

    data_start = HDR_ROW + 1
    for i, row in enumerate(result["schedule"]):
        r = data_start + i
        is_chemo = row["IsChemo"]
        is_addon = row.get("IsAddOn", False)
        warning_text = row.get("Warnings", "") or ""
        has_warning = bool(warning_text)
        if is_chemo:
            consultant_label = "CHEMOASSESSMENT"
        elif is_addon:
            # Preserve the absent consultant's name so the chief can see whose slot.
            consultant_label = f"(ADD-ON) {row['Consultant'].upper()}"
        else:
            consultant_label = row["Consultant"].upper()
        cells = [
            row["Day"],
            date_map.get(row["Day"], ""),
            row["Session"],
            code_for(row["Consultant"]),
            consultant_label,
            row["Specialty"],
            int(row["Patients"]),
            int(row.get("Required", 0)),
            helpers_str(row["Assigned"]),
            warning_text,
        ]
        for col, val in enumerate(cells, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = BOLD if col in (1, 4) else NORM
            c.alignment = LEFT if col in (9, 10) else CENTER
            c.border = BORDER
            if is_chemo:
                c.fill = CHEMO
            elif is_addon:
                # Warning overrides add-on (red beats orange for visual urgency).
                c.fill = WARN_FILL if has_warning else ADDON
            elif has_warning:
                c.fill = WARN_FILL

    DATA_END = data_start + len(result["schedule"]) - 1
    # Adjusted column widths for new layout
    widths = [6, 8, 8, 7, 18, 12, 9, 9, 32, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    # Helpers column moved from H to I in new layout
    helpers_range = f"Schedule!$I${data_start}:$I${DATA_END}"
    spec_range    = f"Schedule!$F${data_start}:$F${DATA_END}"
    cons_range    = f"Schedule!$E${data_start}:$E${DATA_END}"  # used by Assistant_Summary chemo count

    # ============ SHEET 2: FELLOW_SUMMARY ============
    fs = wb.create_sheet("Fellow_Summary")
    fs["A1"] = "FELLOW WORKLOAD (auto-updates from Schedule sheet)"
    fs["A1"].font = TITLE; fs["A1"].alignment = CENTER
    fs.merge_cells("A1:G1"); fs.row_dimensions[1].height = 22

    fs_headers = ["Fellow","Rotation","Role","Total Clinics","Inside Rotation",
                  "Outside Rotation","Rotation %"]
    for col, h in enumerate(fs_headers, start=1):
        c = fs.cell(row=3, column=col, value=h)
        c.font = WHITE; c.fill = SUMHDR; c.alignment = CENTER; c.border = BORDER

    for idx, row in fellows_df.iterrows():
        r = 4 + idx
        role = "Fellow"
        if "Role" in fellows_df.columns and pd.notna(row.get("Role")):
            role = str(row["Role"]).strip().capitalize()
        fs.cell(row=r, column=1, value=str(row["Fellow"]).strip()).font = NORM
        fs.cell(row=r, column=2, value=str(row["Rotation"]).strip()).font = NORM
        fs.cell(row=r, column=3, value=role).font = NORM
        fs.cell(row=r, column=4,
                value=f'=SUMPRODUCT(--ISNUMBER(SEARCH(UPPER(A{r}),{helpers_range})))'
               ).font = BOLD
        fs.cell(row=r, column=5,
                value=f'=SUMPRODUCT(--ISNUMBER(SEARCH(UPPER(A{r}),{helpers_range})),'
                      f'--({spec_range}=B{r}))')
        fs.cell(row=r, column=6, value=f"=D{r}-E{r}")
        fs.cell(row=r, column=7,
                value=f'=IF(D{r}=0,"-",ROUND(E{r}/D{r}*100,0)&"%")')
        for col in range(1, 8):
            cell = fs.cell(row=r, column=col)
            cell.alignment = LEFT if col == 1 else CENTER
            cell.border = BORDER

    tr = 4 + len(fellows_df)
    fs.cell(row=tr, column=1, value="TOTAL").font = BOLD
    for col_letter, col_idx in [("D", 4), ("E", 5), ("F", 6)]:
        c = fs.cell(row=tr, column=col_idx,
                    value=f"=SUM({col_letter}4:{col_letter}{tr-1})")
        c.font = BOLD; c.alignment = CENTER; c.border = BORDER
    for col in range(1, 8):
        fs.cell(row=tr, column=col).fill = GREY
        fs.cell(row=tr, column=col).border = BORDER

    fs_widths = [16, 14, 10, 14, 16, 16, 13]
    for i, w in enumerate(fs_widths, start=1):
        fs.column_dimensions[get_column_letter(i)].width = w

    # ============ SHEET 3: RESIDENT_SUMMARY ============
    # Build per-Type lists once; reuse below for Assistant_Summary.
    def _row_type(row):
        if "Type" in residents_df.columns and pd.notna(row.get("Type")):
            return str(row["Type"]).strip().capitalize()
        return "Resident"

    resident_rows  = [r for _, r in residents_df.iterrows() if _row_type(r) != "Assistant"]
    assistant_rows = [r for _, r in residents_df.iterrows() if _row_type(r) == "Assistant"]

    rs = wb.create_sheet("Resident_Summary")
    rs["A1"] = "RESIDENT WORKLOAD (auto-updates)"
    rs["A1"].font = TITLE; rs["A1"].alignment = CENTER
    rs.merge_cells("A1:E1"); rs.row_dimensions[1].height = 22

    rs_headers = ["Resident", "Min", "Max", "Total Clinics", "Utilization (% of Max)"]
    for col, h in enumerate(rs_headers, start=1):
        c = rs.cell(row=3, column=col, value=h)
        c.font = WHITE; c.fill = SUMHDR; c.alignment = CENTER; c.border = BORDER

    for i, row in enumerate(resident_rows):
        r = 4 + i
        min_v = int(row["Min Clinics"]) if pd.notna(row.get("Min Clinics")) else 0
        max_v = int(row["Max Clinics"]) if pd.notna(row.get("Max Clinics")) else 0
        rs.cell(row=r, column=1, value=str(row["Resident"]).strip()).font = NORM
        rs.cell(row=r, column=2, value=min_v).font = NORM
        rs.cell(row=r, column=3, value=max_v).font = NORM
        rs.cell(row=r, column=4,
                value=f'=SUMPRODUCT(--ISNUMBER(SEARCH(UPPER(A{r}),{helpers_range})))'
               ).font = BOLD
        rs.cell(row=r, column=5,
                value=f'=IF(C{r}=0,"-",ROUND(D{r}/C{r}*100,0)&"%")')
        for col in range(1, 6):
            c = rs.cell(row=r, column=col)
            c.alignment = LEFT if col == 1 else CENTER
            c.border = BORDER

    if resident_rows:
        rtr = 4 + len(resident_rows)
        rs.cell(row=rtr, column=1, value="TOTAL").font = BOLD
        rs.cell(row=rtr, column=4, value=f"=SUM(D4:D{rtr-1})").font = BOLD
        rs.cell(row=rtr, column=4).alignment = CENTER
        for col in range(1, 6):
            rs.cell(row=rtr, column=col).fill = GREY
            rs.cell(row=rtr, column=col).border = BORDER

    for i, w in enumerate([16, 6, 6, 14, 22], start=1):
        rs.column_dimensions[get_column_letter(i)].width = w

    # ============ SHEET 3b: ASSISTANT_SUMMARY (only if any) ============
    # Assistants are chemo-eligible (unlike regular residents). They get their own
    # sheet with an extra Chemo Clinics column so the chief can see at a glance
    # whether assistants are being routed to chemo as intended.
    if assistant_rows:
        ass = wb.create_sheet("Assistant_Summary")
        ass["A1"] = "ASSISTANT WORKLOAD (auto-updates)"
        ass["A1"].font = TITLE; ass["A1"].alignment = CENTER
        ass.merge_cells("A1:F1"); ass.row_dimensions[1].height = 22

        ass_headers = ["Assistant", "Min", "Max", "Total Clinics",
                       "Chemo Clinics", "Utilization (% of Max)"]
        for col, h in enumerate(ass_headers, start=1):
            c = ass.cell(row=3, column=col, value=h)
            c.font = WHITE; c.fill = SUMHDR; c.alignment = CENTER; c.border = BORDER

        for i, row in enumerate(assistant_rows):
            r = 4 + i
            min_v = int(row["Min Clinics"]) if pd.notna(row.get("Min Clinics")) else 0
            max_v = int(row["Max Clinics"]) if pd.notna(row.get("Max Clinics")) else 0
            ass.cell(row=r, column=1, value=str(row["Resident"]).strip()).font = NORM
            ass.cell(row=r, column=2, value=min_v).font = NORM
            ass.cell(row=r, column=3, value=max_v).font = NORM
            ass.cell(row=r, column=4,
                     value=f'=SUMPRODUCT(--ISNUMBER(SEARCH(UPPER(A{r}),{helpers_range})))'
                    ).font = BOLD
            ass.cell(row=r, column=5,
                     value=f'=SUMPRODUCT(--ISNUMBER(SEARCH(UPPER(A{r}),{helpers_range})),'
                           f'--({cons_range}="CHEMOASSESSMENT"))'
                    ).font = BOLD
            ass.cell(row=r, column=6,
                     value=f'=IF(C{r}=0,"-",ROUND(D{r}/C{r}*100,0)&"%")')
            for col in range(1, 7):
                c = ass.cell(row=r, column=col)
                c.alignment = LEFT if col == 1 else CENTER
                c.border = BORDER

        atr = 4 + len(assistant_rows)
        ass.cell(row=atr, column=1, value="TOTAL").font = BOLD
        for col_letter, col_idx in [("D", 4), ("E", 5)]:
            c = ass.cell(row=atr, column=col_idx,
                         value=f"=SUM({col_letter}4:{col_letter}{atr-1})")
            c.font = BOLD; c.alignment = CENTER
        for col in range(1, 7):
            ass.cell(row=atr, column=col).fill = GREY
            ass.cell(row=atr, column=col).border = BORDER

        for i, w in enumerate([16, 6, 6, 14, 14, 22], start=1):
            ass.column_dimensions[get_column_letter(i)].width = w

    # ============ SHEET 4: NOTES ============
    nt = wb.create_sheet("Notes")
    nt["A1"] = "LEGEND & NOTES"
    nt["A1"].font = TITLE
    nt.merge_cells("A1:B1")
    nt.row_dimensions[1].height = 22

    notes = [
        ("Code", "Meaning"),
        ("CA", "Chemoassessment Clinic"),
        ("CB", "Combined Clinic — varies weekly, fill in manually"),
        ("*",  "External staff covering via purple rule"),
        ("ADD-ON", "Consultant on leave; ~8-10 pt slot covered solo by fellow (rotation match) or assistant"),
        ("", ""),
        ("Workflow", ""),
        ("1.", "Edit any cell in the 'Helpers' column on Schedule."),
        ("2.", "Use '/' to separate multiple helpers (e.g. ALHARBI/FATIMAH)."),
        ("3.", "Fellow_Summary, Resident_Summary, and Assistant_Summary recalculate automatically."),
        ("4.", "Names are case-insensitive."),
        ("5.", "To mark a clinic as add-on, set 'ClinicType' column to 'AddOn' in clinics.xlsx."),
        ("", ""),
        ("Helper Rule", "Helpers by patient volume (regular clinics only — add-ons always need 1)"),
        ("0-11", "0 helpers"),
        ("12-14", "1 (resident/NP preferred)"),
        ("15-16", "1 fellow OR 2 anything"),
        ("17-19", "2 helpers (1 fellow + 1 non-fellow preferred)"),
        ("20-23", "2 fellows alone, OR 1 fellow + 2 non-fellows"),
        ("24-26", "3 helpers (2 fellows + 1 non-fellow)"),
        ("27+",   "3 fellows OR 2 fellows + 2 non-fellows"),
    ]
    for i, (a, b) in enumerate(notes, start=3):
        nt.cell(row=i, column=1, value=a).font = BOLD if a in ("Code","Workflow","Helper Rule") else NORM
        nt.cell(row=i, column=2, value=b).font = BOLD if a in ("Code","Workflow","Helper Rule") else NORM
        if a in ("Code", "Workflow", "Helper Rule"):
            nt.cell(row=i, column=1).fill = GREY
            nt.cell(row=i, column=2).fill = GREY

    nt.column_dimensions["A"].width = 12
    nt.column_dimensions["B"].width = 80

    # ============ v12.2: On-Leave + Availability Orphans + Purple Orphans summary ============
    # Surface "who's off this week" and any stale rows pointing at on-Leave or unknown names.
    on_leave_list = result.get("on_leave", [])
    orphan_list = result.get("availability_orphans", [])
    purple_stale = result.get("purple_orphans_leave", [])
    if on_leave_list or orphan_list or purple_stale:
        row_start = len(notes) + 5  # leave a blank line after the legend
        nt.cell(row=row_start, column=1, value="Roster Notes").font = BOLD
        nt.cell(row=row_start, column=1).fill = GREY
        nt.cell(row=row_start, column=2, value="").fill = GREY
        r = row_start + 1
        if on_leave_list:
            nt.cell(row=r, column=1, value="On Leave").font = BOLD
            nt.cell(row=r, column=2, value=", ".join(on_leave_list)).font = NORM
            r += 1
        if orphan_list:
            nt.cell(row=r, column=1, value="Avail. skipped").font = BOLD
            orphans_text = "; ".join(
                f"{o['person']} ({o['day']} {o['session']})" for o in orphan_list
            )
            nt.cell(row=r, column=2,
                    value=f"Not in this week's roster: {orphans_text}").font = NORM
            r += 1
        if purple_stale:
            nt.cell(row=r, column=1, value="Purple skipped").font = BOLD
            stale_text = "; ".join(
                f"{p['person']} ({p['day'].title()} {p['session']} {p['consultant'].title()})"
                for p in purple_stale
            )
            nt.cell(row=r, column=2,
                    value=f"On Leave this week: {stale_text}").font = NORM
            r += 1

    # ============ SHEET: LEAVE_ENTRIES (only if any) ============
    avail = result.get("availability", [])
    if avail:
        lv = wb.create_sheet("Leave_Entries")
        lv["A1"] = "LEAVE / AVAILABILITY ENTRIES APPLIED"
        lv["A1"].font = TITLE
        lv.merge_cells("A1:D1")
        lv.row_dimensions[1].height = 22

        for col, h in enumerate(["Person", "Day", "Session", "Reason"], start=1):
            c = lv.cell(row=3, column=col, value=h)
            c.font = WHITE; c.fill = SUMHDR; c.alignment = CENTER; c.border = BORDER

        for idx, entry in enumerate(avail, start=4):
            lv.cell(row=idx, column=1, value=entry["person"]).font = NORM
            lv.cell(row=idx, column=2, value=entry["day"].capitalize()).font = NORM
            lv.cell(row=idx, column=3, value=entry["session"]).font = NORM
            lv.cell(row=idx, column=4, value=entry.get("reason", "")).font = NORM
            for col in range(1, 5):
                cell = lv.cell(row=idx, column=col)
                cell.alignment = LEFT if col in (1, 4) else CENTER
                cell.border = BORDER

        for col, w in enumerate([16, 12, 10, 24], start=1):
            lv.column_dimensions[get_column_letter(col)].width = w

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
